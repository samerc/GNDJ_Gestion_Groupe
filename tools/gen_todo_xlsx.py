# -*- coding: utf-8 -*-
"""Extract EVERY checklist item ([x] / [ ]) from CLAUDE.md into an editable Excel task tracker."""
import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = r"C:\Users\Administrator\Documents\coding\GNDJ_Gestion_Groupe\CLAUDE.md"
OUT = r"C:\Users\Administrator\Desktop\GNDJ_taches.xlsx"

def clean(s: str) -> str:
    s = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", s)   # [text](url) -> text
    s = s.replace("**", "").replace("`", "")
    s = re.sub(r"~~([^~]*)~~", r"\1", s)               # strike -> text
    s = re.sub(r"\s+", " ", s).strip()
    return s

lines = open(SRC, encoding="utf-8").read().splitlines()
task_re = re.compile(r"^(\s*)-\s*\[([ xX])\]\s*(.*)$")
head_re = re.compile(r"^#{2,4}\s+(.*)$")
bullet_re = re.compile(r"^\s*-\s")
heading_line_re = re.compile(r"^#")

section = "(intro)"
rows = []  # (section, status, text)
i = 0
while i < len(lines):
    ln = lines[i]
    h = head_re.match(ln)
    if h:
        section = clean(h.group(1))
        i += 1
        continue
    m = task_re.match(ln)
    if m:
        done = m.group(2).lower() == "x"
        text = m.group(3)
        # join wrapped continuation lines (indented, not a new bullet/heading/blank)
        j = i + 1
        while j < len(lines):
            nxt = lines[j]
            if nxt.strip() == "" or bullet_re.match(nxt) or heading_line_re.match(nxt):
                break
            text += " " + nxt.strip()
            j += 1
        rows.append((section, "Fait" if done else "À faire", clean(text)))
        i = j
        continue
    i += 1

# ---- build workbook ----
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tâches GNDJ"
headers = ["#", "Section / Phase", "Tâche", "Statut d'origine", "Nouveau statut", "Priorité", "Notes"]

navy = "1F3A5F"; light = "EAF1F6"
thin = Side(style="thin", color="C9D4DE"); border = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="top", wrap_text=True)

ws.merge_cells("A1:G1")
ws["A1"] = f"GNDJ — Toutes les tâches extraites de CLAUDE.md ({len(rows)} lignes, 2026-08-25)"
ws["A1"].font = Font(bold=True, size=13, color=navy)
ws.row_dimensions[1].height = 24

for c, hh in enumerate(headers, 1):
    cell = ws.cell(2, c, hh)
    cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cell.border = border
ws.row_dimensions[2].height = 22

done_fill = PatternFill("solid", fgColor="D4EDDA")
todo_fill = PatternFill("solid", fgColor="FFF3CD")
zebra = PatternFill("solid", fgColor=light)

r = 3
for idx, (sec, status, text) in enumerate(rows, 1):
    vals = [idx, sec, text, status, "", "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.border = border
        cell.alignment = ctr if c in (1, 4, 5, 6) else wrap
    if idx % 2 == 0:
        for c in (1, 2, 3, 7):
            ws.cell(r, c).fill = zebra
    ws.cell(r, 4).fill = done_fill if status == "Fait" else todo_fill
    r += 1

for col, w in {"A":5, "B":34, "C":80, "D":15, "E":15, "F":12, "G":30}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:G{r-1}"

stat_dv = DataValidation(type="list", formula1='"À faire,En cours,Bloqué,Fait,À refaire,Abandonné,À revoir"', allow_blank=True)
prio_dv = DataValidation(type="list", formula1='"Haute,Moyenne,Basse,Optionnelle"', allow_blank=True)
ws.add_data_validation(stat_dv); ws.add_data_validation(prio_dv)
stat_dv.add(f"E3:E{r-1}"); prio_dv.add(f"F3:F{r-1}")

# Summary sheet: counts per section
ws2 = wb.create_sheet("Résumé par section")
ws2.append(["Section / Phase", "Fait", "À faire", "Total"])
for c in range(1,5):
    cell = ws2.cell(1, c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
from collections import OrderedDict
agg = OrderedDict()
for sec, status, _ in rows:
    d = agg.setdefault(sec, [0,0]); d[0 if status=="Fait" else 1] += 1
for sec,(done,todo) in agg.items():
    ws2.append([sec, done, todo, done+todo])
for col,w in {"A":50,"B":8,"C":10,"D":8}.items(): ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"

wb.save(OUT)
done_n = sum(1 for _,s,_ in rows if s=="Fait")
print(f"Saved {OUT}: {len(rows)} tasks ({done_n} Fait, {len(rows)-done_n} À faire), {len(agg)} sections")
