# -*- coding: utf-8 -*-
"""Build an editable Excel from the Rentrée checklist, collapsing per-unit fan-out tasks into ONE row
(with an X/N units progress), the way the app's Rentrée page rolls them up."""
import csv, openpyxl
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

CSV = r"C:\tmp\rentree2.csv"
OUT = r"C:\Users\Administrator\Desktop\GNDJ_rentree_taches.xlsx"

ROLE = {
    "chef-de-groupe": "Chef de Groupe", "assistant-de-groupe": "Assistant de Groupe",
    "chef-unite": "Chef d'unité", "chef-equipe": "Chef d'équipe",
    "association-admin": "Admin association", "read-only": "Membre",
}
ANCHOR = {
    "passage.date": "Date du passage", "demande.submission_start": "Ouverture des soumissions",
    "demande.submission_deadline": "Date limite des soumissions",
    "demande.member_start_date": "Début des nouveaux membres",
    "documents.deposit_start": "Ouverture dépôt documents",
    "documents.verification1_start": "Vérification 1", "documents.correction_start": "Correction",
    "documents.verification2_start": "Vérification 2", "documents.final_deadline": "Date limite finale",
}
rows = list(csv.DictReader(open(CSV, encoding="utf-8")))

# --- group: per-unit fan-out tasks share a template_id -> collapse into one group; keep display order ---
groups = OrderedDict()  # key -> {meta + units list}
for r in rows:
    key = r["template_id"] or ("one:" + r["id"])   # null-template one-offs stay individual
    g = groups.get(key)
    if g is None:
        g = groups[key] = {
            "phase": r["phase"], "order": int(r["display_order"]), "title": r["title"],
            "description": r["description"], "assignee_role": r["assignee_role"],
            "deadline_label": r["deadline_label"], "due_date": r["due_date"],
            "deadline_anchor": r["deadline_anchor"], "progress_key": r["progress_key"],
            "action_key": r["action_key"], "units": [], "done": 0, "total": 0,
        }
    g["total"] += 1
    if r["status"] == "done":
        g["done"] += 1
    if r["unit_name"]:
        g["units"].append(r["unit_name"])

items = sorted(groups.values(), key=lambda g: (g["order"], g["title"]))

def echeance(g):
    base = g["deadline_label"] or g["due_date"] or "—"
    if g["deadline_anchor"]:
        base += f"  (auto : {ANCHOR.get(g['deadline_anchor'], g['deadline_anchor'])})"
    return base

def scope(g):
    n = len(g["units"])
    return f"Par unité — {n} unités" if n else "Groupe"

def progress(g):
    if g["units"]:
        return f"{g['done']}/{g['total']} unités faites"
    return "Fait" if g["done"] == g["total"] and g["total"] else "À faire"

def statut(g):
    if g["done"] == 0:
        return "À faire"
    if g["done"] == g["total"]:
        return "Fait"
    return "Partiel"

# --- workbook ---
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Rentrée 2026-2027"
navy = "1F3A5F"; light = "EAF1F6"
thin = Side(style="thin", color="C9D4DE"); border = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical="top"); ctr = Alignment("center", "top", wrap_text=True)

headers = ["#", "Phase", "Tâche", "Description", "Portée", "Responsable",
           "Échéance", "Suivi auto.", "Action", "Avancement", "Statut d'origine", "Nouveau statut", "Notes"]

done_full = sum(1 for g in items if g["done"] == g["total"])
partial = sum(1 for g in items if 0 < g["done"] < g["total"])
ws.merge_cells("A1:M1")
ws["A1"] = (f"GNDJ — Rentrée scoute 2026-2027 : liste de tâches (regroupée) — "
            f"{len(items)} tâches ({done_full} faites, {partial} partielles, {len(items)-done_full-partial} à faire)")
ws["A1"].font = Font(bold=True, size=13, color=navy); ws.row_dimensions[1].height = 24

for c, h in enumerate(headers, 1):
    cell = ws.cell(2, c, h)
    cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); cell.border = border
ws.row_dimensions[2].height = 22

done_fill = PatternFill("solid", fgColor="D4EDDA"); todo_fill = PatternFill("solid", fgColor="FFF3CD")
part_fill = PatternFill("solid", fgColor="FDE9D0"); zebra = PatternFill("solid", fgColor=light)

rr = 3
for i, g in enumerate(items, 1):
    resp = ROLE.get(g["assignee_role"], g["assignee_role"] or "—")
    suivi = "auto : " + g["progress_key"] if g["progress_key"] else ""
    st = statut(g)
    vals = [i, g["phase"], g["title"], g["description"], scope(g), resp,
            echeance(g), suivi, g["action_key"] or "", progress(g), st, "", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(rr, c, v); cell.border = border
        cell.alignment = ctr if c in (1, 10, 11, 12) else wrap
    if i % 2 == 0:
        for c in (1, 2, 3, 4, 5, 6, 7, 8, 9, 13): ws.cell(rr, c).fill = zebra
    ws.cell(rr, 11).fill = {"Fait": done_fill, "Partiel": part_fill, "À faire": todo_fill}[st]
    rr += 1

for col, w in {"A":5,"B":16,"C":42,"D":54,"E":18,"F":20,"G":26,"H":20,"I":18,"J":18,"K":15,"L":15,"M":26}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:M{rr-1}"

dv = DataValidation(type="list", formula1='"À faire,En cours,Bloqué,Partiel,Fait,À revoir,Sans objet"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"L3:L{rr-1}")

# Summary per phase (grouped counts)
ws2 = wb.create_sheet("Résumé par phase")
ws2.append(["Phase", "Faites", "Partielles", "À faire", "Total"])
for c in range(1, 6):
    cell = ws2.cell(1, c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
agg = OrderedDict()
for g in items:
    a = agg.setdefault(g["phase"], [0, 0, 0])
    st = statut(g); a[0 if st == "Fait" else 1 if st == "Partiel" else 2] += 1
for ph, (d, p, t) in agg.items(): ws2.append([ph, d, p, t, d + p + t])
ws2.append(["TOTAL", done_full, partial, len(items) - done_full - partial, len(items)])
for col, w in {"A":22, "B":8, "C":11, "D":8, "E":8}.items(): ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved {OUT}: {len(items)} grouped tasks (from {len(rows)} rows), {len(agg)} phases")
